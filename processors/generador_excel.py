# -*- coding: utf-8 -*-
"""
generador_excel.py

Genera el Excel siguiendo las plantillas oficiales del portal EGEHID.
Existen 3 plantillas distintas, una por cada tipo de documento:

    "nota_credito"  → 31 columnas (A-AE)
    "nota_debito"   → 36 columnas (A-AJ)  — agrega datos de pago/banco
    "factura"       → 68 columnas (A-BP)  — agrega descuentos, retenciones
                       ITBIS/ISR y datos completos de embarque/transporte

Todas las columnas de la plantilla correspondiente se escriben SIEMPRE,
aunque estén vacías (celda presente con valor None), para respetar
exactamente la estructura que exige el portal.

Fila 1 – encabezados para filas E (encabezado de documento)
Fila 2 – encabezados para filas D (detalle de ítem)
"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ── Estilos ───────────────────────────────────────────────────────────────────
# Encabezados en gris (antes rojo). GRIS_HDR = gris medio para las columnas
# que existen en ambas filas (E y D); GRIS_HDR_OSC = gris oscuro para las
# columnas que solo aplican a la fila E.
GRIS_HDR     = PatternFill("solid", start_color="595959", end_color="595959")
GRIS_HDR_OSC = PatternFill("solid", start_color="333333", end_color="333333")
AZUL_E     = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
FUENTE_HDR = Font(bold=True, color="FFFFFF", size=10)
FUENTE_E   = Font(bold=True, size=10)
FUENTE_D   = Font(size=10)

FORMATO_FECHA = "DD/MM/YYYY"
FORMATO_NUM   = "#,##0.00"

NOMBRE_HOJA = "Facturacion"

# Palabras clave (en el encabezado de la columna) que identifican una columna
# MONETARIA. Sirve para dar a los importes el formato numérico de moneda
# (#,##0.00), sin depender de la posición ni de la plantilla concreta.
_PALABRAS_MONETARIAS = ("precio", "monto", "valor descuento", "tasa cambio")


def _es_columna_monetaria(encabezado) -> bool:
    """True si el encabezado de la columna corresponde a un importe."""
    if not encabezado:
        return False
    h = str(encabezado).lower()
    return any(p in h for p in _PALABRAS_MONETARIAS)

# Fecha de expiración del NCF (autorización e-CF de la DGII). Es un valor
# FIJO que no aparece en el PDF; el portal EGEHID lo exige en la columna
# "Fecha Expiracion NCF" de las plantillas de Factura y Nota de Débito.
# Actualízalo cuando la DGII renueve la vigencia de la secuencia.
FECHA_EXPIRACION_NCF = datetime(2027, 12, 31)


def _a_entero(valor):
    """Convierte un texto puramente numérico a int (sin ceros a la izquierda);
    en cualquier otro caso devuelve el valor tal cual. El portal espera
    'Numero Factura Interna' y 'RNC Comprador' como números, no como texto."""
    if isinstance(valor, str):
        v = valor.strip()
        if v.isdigit() and (v == "0" or not v.startswith("0")):
            return int(v)
    return valor


# ════════════════════════════════════════════════════════════════════════════
# DEFINICIÓN DE LAS 3 PLANTILLAS
# ════════════════════════════════════════════════════════════════════════════
# Cada entrada de COLS: (col, clave_interna, header_fila1_E, header_fila2_D, ancho)
# header_fila2_D = None  → la columna no existe en la fila D (detalle);
#                           la celda se deja vacía y el header de fila 2
#                           se pinta gris en vez de rojo.

# ════════════════════════════════════════════════════════════════════════════
# CONVENCIÓN DE CLAVES PARA FILAS D (DETALLE)
# ════════════════════════════════════════════════════════════════════════════
# Cada plantilla tiene sus propios campos de detalle en posiciones distintas.
# Para evitar desplazamientos usamos claves semánticamente exactas que
# corresponden 1-a-1 con el campo real de la columna en esa plantilla.
#
# Claves compartidas entre las 3 plantillas (filas E y D):
#   tipo_linea, ncf → siempre en cols 1 y 2
#
# Claves de detalle (fila D) — nombradas según el campo real del portal:
#   d_nombre_item          → "Nombre Item"
#   d_desc_item            → "Descripción Item"
#   d_ind_bien_servicio    → "Indicador Bien o Servicio"     (siempre 2)
#   d_ind_tasa_itbis       → "Indicador Tasa ITBIS"          (siempre 4)
#   d_cantidad             → "Cantidad Item"                  (del PDF)
#   d_precio_unitario      → "Precio Unitario Item"           (del PDF)
#   d_unidad_medida        → "Unidad de Medida"               (del PDF, o None)
# ════════════════════════════════════════════════════════════════════════════

# ── Plantilla: NOTA DE CRÉDITO (31 columnas, A-AE) ───────────────────────────
# Fila D: col3=Nombre Item, col4=Desc Item, col5=Ind Bien/Serv,
#         col6=Ind Tasa ITBIS, col7=Cantidad, col8=Precio Unitario,
#         col9=Unidad de Medida
COLS_NOTA_CREDITO = [
    ( 1,  "tipo_linea",          "Tipo de Linea",                    "Tipo de Linea",               14),
    ( 2,  "ncf",                 "NCF",                              "NCF",                         22),
    ( 3,  "fecha_documento",     "Fecha del Documento",              "Nombre Item",                 24),
    ( 4,  "numero_documento",    "Número Factura Interna",           "Descripción Item",            22),
    ( 5,  "ind_ingreso",         "Tipo de Ingresos",                 "Indicador Bien o Servicio",   20),
    ( 6,  "tipo_pago",           "Tipo de Pago",                     "Indicador Tasa ITBIS",        18),
    ( 7,  "fecha_vencimiento",   "Fecha Limite Pago",                "Cantidad Item",               20),
    ( 8,  "moneda",              "Moneda",                           "Precio Unitario Item",        16),
    ( 9,  "tasa_cambio",         "Tasa Cambio",                      "Unidad de Medida",            14),
    (10,  "rnc_cliente",         "RNC Comprador",                    "Fecha Vencimiento Item",      18),
    (11,  "cliente",             "Razon Social Comprador",           "Fecha de Elaboración",        30),
    (12,  "contacto_comprador",  "Contacto del Comprador",           "Grados Alcohol",              20),
    (13,  "correo_comprador",    "Correo del Comprador",             "Unidad de Referencia",        24),
    (14,  "direccion_comprador", "Direccion del Comprador",          "Cantidad de Referencia",      24),
    (15,  "municipio_comprador", "Municipio del Comprador",          "Tipo de Codigo 1",            20),
    (16,  "provincia_comprador", "Provincia del Comprador",          "Codigo Item 1",               20),
    (17,  "fecha_entrega",       "Fecha de Entrega",                 "Tipo de Codigo 2",            18),
    (18,  "contacto_entrega",    "Contacto de Entrega",              "Codigo Item 2",               20),
    (19,  "direccion_entrega",   "Dirección de Entrega",             "Tipo de Codigo 3",            24),
    (20,  "telefono_adicional",  "Telefono Adicional",               "Codigo Item 3",               18),
    (21,  "fecha_orden_compra",  "Fecha de Orden Compra",            "Tipo de Codigo 4",            20),
    (22,  "numero_orden_compra", "Numero de Orden Compra",           "Codigo Item 4",               20),
    (23,  "codigo_interno",      "Codigo Interno del Comprador",     "Tipo de Codigo 5",            24),
    (24,  "responsable_pago",    "Responsable de Pago",              "Codigo Item 5",               20),
    (25,  "info_adicional",      "Información Adicional Comprador",  None,                          24),
    (26,  "ncf_modificado",      "NCF Modificado",                   None,                          22),
    (27,  "rnc_otro",            "RNC Otro Contribuyente",           None,                          20),
    (28,  "fecha_ncf_modificado","Fecha NCF Modificado",             None,                          20),
    (29,  "cod_modificacion",    "Codigo de Modificación",           None,                          18),
    (30,  "razon_modificacion",  "Razon de Modificación",            None,                          20),
    (31,  "nota",                "Nota",                             None,                          60),
]
# Mapeo D para Nota de Crédito: clave_interna → clave_detalle
# col3=fecha_documento→d_nombre_item, col4=numero_documento→d_desc_item,
# col5=ind_ingreso→d_ind_bien_servicio, col6=tipo_pago→d_ind_tasa_itbis,
# col7=fecha_vencimiento→d_cantidad, col8=moneda→d_precio_unitario,
# col9=tasa_cambio→d_unidad_medida
D_MAP_NOTA_CREDITO = {
    "fecha_documento":  "d_nombre_item",
    "numero_documento": "d_desc_item",
    "ind_ingreso":      "d_ind_bien_servicio",
    "tipo_pago":        "d_ind_tasa_itbis",
    "fecha_vencimiento":"d_cantidad",
    "moneda":           "d_precio_unitario",
    "tasa_cambio":      "d_unidad_medida",
}

# ── Plantilla: NOTA DE DÉBITO (36 columnas, A-AJ) ────────────────────────────
# Fila D: col3=Nombre Item, col4=Desc Item, col5=Ind Bien/Serv,
#         col6=Ind Tasa ITBIS, col7=Cantidad, col8=Precio Unitario,
#         col9=Unidad de Medida
COLS_NOTA_DEBITO = [
    ( 1,  "tipo_linea",          "Tipo de Linea",                    "Tipo de Linea",               14),
    ( 2,  "ncf",                 "NCF",                              "NCF",                         22),
    ( 3,  "fecha_expiracion_ncf","Fecha Expiración NCF",             "Nombre Item",                 22),
    ( 4,  "fecha_documento",     "Fecha del Documento",              "Descripción Item",            24),
    ( 5,  "numero_documento",    "Numero Factura Interna",           "Indicador Bien o Servicio",   22),
    ( 6,  "ind_ingreso",         "Tipo de Ingresos",                 "Indicador Tasa ITBIS",        20),
    ( 7,  "tipo_pago",           "Tipo de Pago",                     "Cantidad Item",               18),
    ( 8,  "termino_pago",        "Termino de Pago",                  "Precio Unitario Item",        18),
    ( 9,  "fecha_vencimiento",   "Fecha Limite Pago",                "Unidad de Medida",            20),
    (10,  "tipo_cuenta_pago",    "Tipo Cuenta de Pago",              "Fecha Vencimiento Item",      18),
    (11,  "numero_cuenta_pago",  "Numero de Cuenta Pago",            "Fecha de Elaboración",        20),
    (12,  "banco_pago",          "Banco de Pago",                    "Grados Alcohol",              18),
    (13,  "moneda",              "Moneda",                           "Unidad de Referencia",        16),
    (14,  "tasa_cambio",         "Tasa Cambio",                      "Cantidad de Referencia",      14),
    (15,  "rnc_cliente",         "RNC Comprador",                    "Tipo de Codigo 1",            18),
    (16,  "cliente",             "Razon Social Comprador",           "Codigo Item 1",               30),
    (17,  "contacto_comprador",  "Contacto del Comprador",           "Tipo de Codigo 2",            20),
    (18,  "correo_comprador",    "Correo del Comprador",             "Codigo Item 2",               24),
    (19,  "direccion_comprador", "Direccion del Comprador",          "Tipo de Codigo 3",            24),
    (20,  "municipio_comprador", "Municipio del Comprador",          "Codigo Item 3",               20),
    (21,  "provincia_comprador", "Provincia del Comprador",          "Tipo de Codigo 4",            20),
    (22,  "fecha_entrega",       "Fecha de Entrega",                 "Codigo Item 4",               18),
    (23,  "contacto_entrega",    "Contacto de Entrega",              "Tipo de Codigo 5",            20),
    (24,  "direccion_entrega",   "Dirección de Entrega",             "Codigo Item 5",               24),
    (25,  "telefono_adicional",  "Telefono Adicional",               None,                          18),
    (26,  "fecha_orden_compra",  "Fecha de Orden Compra",            None,                          20),
    (27,  "numero_orden_compra", "Numero de Orden Compra",           None,                          20),
    (28,  "codigo_interno",      "Codigo Interno del Comprador",     None,                          24),
    (29,  "responsable_pago",    "Responsable de Pago",              None,                          20),
    (30,  "info_adicional",      "Informacion Adicional Comprador",  None,                          24),
    (31,  "ncf_modificado",      "NCF Modificado",                   None,                          22),
    (32,  "rnc_otro",            "RNC Otro Contribuyente",           None,                          20),
    (33,  "fecha_ncf_modificado","Fecha NCF Modificado",             None,                          20),
    (34,  "cod_modificacion",    "Codigo de Modificación",           None,                          18),
    (35,  "razon_modificacion",  "Razon de Modificación",            None,                          20),
    (36,  "nota",                "Nota",                             None,                          60),
]
# Mapeo D para Nota de Débito:
# col3=fecha_expiracion_ncf→d_nombre_item, col4=fecha_documento→d_desc_item,
# col5=numero_documento→d_ind_bien_servicio, col6=ind_ingreso→d_ind_tasa_itbis,
# col7=tipo_pago→d_cantidad, col8=termino_pago→d_precio_unitario,
# col9=fecha_vencimiento→d_unidad_medida
D_MAP_NOTA_DEBITO = {
    "fecha_expiracion_ncf": "d_nombre_item",
    "fecha_documento":      "d_desc_item",
    "numero_documento":     "d_ind_bien_servicio",
    "ind_ingreso":          "d_ind_tasa_itbis",
    "tipo_pago":            "d_cantidad",
    "termino_pago":         "d_precio_unitario",
    "fecha_vencimiento":    "d_unidad_medida",
}

# ── Plantilla: FACTURA (68 columnas, A-BP) ────────────────────────────────────
# Fila D: col3=Nombre Item, col4=Desc Item, col5=Ind Bien/Serv,
#         col6=Ind Tasa ITBIS, col7=Cantidad, col8=Precio Unitario,
#         col9=Unidad de Medida
COLS_FACTURA = [
    ( 1,  "tipo_linea",          "Tipo de Linea",                       "Tipo de Linea",                   14),
    ( 2,  "ncf",                 "NCF",                                 "NCF",                             22),
    ( 3,  "fecha_expiracion_ncf","Fecha Expiracion NCF",                "Nombre Item",                     22),
    ( 4,  "fecha_documento",     "Fecha del Documento",                 "Descripción Item",                24),
    ( 5,  "numero_documento",    "Numero Factura Interna",              "Indicador Bien o Servicio",       22),
    ( 6,  "ind_ingreso",         "Tipo de Ingresos",                    "Indicador Tasa ITBIS",            20),
    ( 7,  "tipo_pago",           "Tipo de Pago",                        "Cantidad Item",                   18),
    ( 8,  "termino_pago",        "Termino de Pago",                     "Precio Unitario Item",            18),
    ( 9,  "fecha_vencimiento",   "Fecha Limite Pago",                   "Unidad de Medida",                20),
    (10,  "tipo_cuenta_pago",    "Tipo Cuenta de Pago",                 "Indicador Retencion o Percepcion",20),
    (11,  "numero_cuenta_pago",  "Numero de Cuenta Pago",               "Monto ITBIS Retenido",            20),
    (12,  "banco_pago",          "Banco de Pago",                       "Monto ISR Retenido",              18),
    (13,  "moneda",              "Moneda",                              "Fecha Vencimiento Item",          16),
    (14,  "tasa_cambio",         "Tasa Cambio",                         "Fecha de Elaboración",            14),
    (15,  "monto_descuento",     "Monto Descuento",                     "Grados Alcohol",                  18),
    (16,  "valor_descuento",     "Valor Descuento",                     "Unidad de Referencia",            18),
    (17,  "tipo_valor",          "Tipo Valor",                          "Cantidad de Referencia",          16),
    (18,  "desc_descuento",      "Descripcion Descuento",               "Tipo de Codigo 1",                22),
    (19,  "ind_norma_1007",      "Indicador Norma 1007",                "Codigo Item 1",                   20),
    (20,  "ind_facturacion_desc","Indicador Facturacion Descuento",     "Tipo de Codigo 2",                24),
    (21,  "rnc_cliente",         "RNC Comprador",                       "Codigo Item 2",                   18),
    (22,  "cliente",             "Razon Social Comprador",              "Tipo de Codigo 3",                30),
    (23,  "contacto_comprador",  "Contacto del Comprador",              "Codigo Item 3",                   20),
    (24,  "correo_comprador",    "Correo del Comprador",                "Tipo de Codigo 4",                24),
    (25,  "direccion_comprador", "Direccion del Comprador",             "Codigo Item 4",                   24),
    (26,  "municipio_comprador", "Municipio del Comprador",             "Tipo de Codigo 5",                20),
    (27,  "provincia_comprador", "Provincia del Comprador",             "Codigo Item 5",                   20),
    (28,  "fecha_entrega",       "Fecha de Entrega",                    "Tipo SubDescuento",               18),
    (29,  "contacto_entrega",    "Contacto de Entrega",                 "SubDescuento Porcentaje",         20),
    (30,  "direccion_entrega",   "Dirección de Entrega",                "Monto SubDescuento",              24),
    (31,  "telefono_adicional",  "Telefono Adicional",                  "Tipo SubRecargo",                 18),
    (32,  "fecha_orden_compra",  "Fecha de Orden Compra",               "SubRecargo Porcentaje",           20),
    (33,  "numero_orden_compra", "Numero de Orden Compra",              "Monto SubRecargo",                20),
    (34,  "codigo_interno",      "Codigo Interno del Comprador",        None,                              24),
    (35,  "responsable_pago",    "Responsable de Pago",                 None,                              20),
    (36,  "info_adicional",      "Informacion Adicional Comprador",     None,                              24),
    (37,  "ncf_modificado",      "NCF Modificado",                      None,                              22),
    (38,  "rnc_otro",            "RNC Otro Contribuyente",              None,                              20),
    (39,  "fecha_ncf_modificado","Fecha NCF Modificado",                None,                              20),
    (40,  "cod_modificacion",    "Codigo de Modificación",              None,                              18),
    (41,  "fecha_embarque",      "Fecha de Embarque",                   None,                              18),
    (42,  "numero_embarque",     "Numero de Embarque",                  None,                              20),
    (43,  "numero_contenedor",   "Numero de Contenedor",                None,                              20),
    (44,  "numero_referencia",   "Numero de Referencia",                None,                              20),
    (45,  "puerto_embarque",     "Nombre de Puerto de Embarque",        None,                              24),
    (46,  "peso_bruto",          "Peso Bruto",                          None,                              14),
    (47,  "peso_neto",           "Peso Neto",                           None,                              14),
    (48,  "unidad_peso_bruto",   "Unidad Peso Bruto",                   None,                              16),
    (49,  "unidad_peso_neto",    "Unidad Peso Neto",                    None,                              16),
    (50,  "cantidad_bulto",      "Cantidad Bulto",                      None,                              16),
    (51,  "unidad_bulto",        "Unidad Bulto",                        None,                              14),
    (52,  "volumen_bulto",       "Volumen Bulto",                       None,                              16),
    (53,  "unidad_volumen",      "Unidad Volumen",                      None,                              16),
    (54,  "via_transporte",      "Via de Transporte",                   None,                              18),
    (55,  "pais_origen",         "Pais de Origen",                      None,                              16),
    (56,  "direccion_destino",   "Direccion Destino",                   None,                              22),
    (57,  "pais_destino",        "Pais Destino",                        None,                              16),
    (58,  "rnc_transportista",   "RNC Identificacion Compania Transportista", None,                        24),
    (59,  "nombre_transportista","Nombre Compania Transportista",       None,                              24),
    (60,  "numero_viaje",        "Numero Viaje",                        None,                              16),
    (61,  "conductor",           "Conductor",                           None,                              18),
    (62,  "documento_transporte","Documento de Transporte",             None,                              22),
    (63,  "ficha",               "Ficha",                               None,                              14),
    (64,  "placa",               "Placa",                               None,                              14),
    (65,  "ruta_transporte",     "Ruta de Transporte",                  None,                              18),
    (66,  "zona_transporte",     "Zona de Transporte",                  None,                              18),
    (67,  "numero_albaran",      "NumeroAlbaran",                       None,                              18),
    (68,  "nota",                "Nota",                                None,                              60),
]
# Mapeo D para Factura — idéntico al de Nota de Débito en las primeras 9 cols:
# col3=fecha_expiracion_ncf→d_nombre_item, col4=fecha_documento→d_desc_item,
# col5=numero_documento→d_ind_bien_servicio, col6=ind_ingreso→d_ind_tasa_itbis,
# col7=tipo_pago→d_cantidad, col8=termino_pago→d_precio_unitario,
# col9=fecha_vencimiento→d_unidad_medida
D_MAP_FACTURA = {
    "fecha_expiracion_ncf": "d_nombre_item",
    "fecha_documento":      "d_desc_item",
    "numero_documento":     "d_ind_bien_servicio",
    "ind_ingreso":          "d_ind_tasa_itbis",
    "tipo_pago":            "d_cantidad",
    "termino_pago":         "d_precio_unitario",
    "fecha_vencimiento":    "d_unidad_medida",
}


# ── Registro de plantillas ────────────────────────────────────────────────────
# Cada entrada: (columnas, mapeo_D)
# mapeo_D: {clave_interna_E → clave_semantica_D}
# La clave semántica D se resuelve en _construir_filas a su valor correcto.
PLANTILLAS = {
    "nota_credito": (COLS_NOTA_CREDITO, D_MAP_NOTA_CREDITO),
    "nota_debito":  (COLS_NOTA_DEBITO,  D_MAP_NOTA_DEBITO),
    "factura":      (COLS_FACTURA,      D_MAP_FACTURA),
}

TIPO_POR_DEFECTO = "nota_credito"


def _resolver_plantilla(tipo_documento: str | None):
    """Devuelve (cols, claves, d_map) para el tipo de documento dado.
    Si el tipo no se reconoce, cae a la plantilla por defecto."""
    tipo = tipo_documento if tipo_documento in PLANTILLAS else TIPO_POR_DEFECTO
    cols, d_map = PLANTILLAS[tipo]
    claves = [c[1] for c in cols]
    return cols, claves, d_map


# ── Lectura de duplicados existentes ─────────────────────────────────────────
def obtener_documentos_existentes(ruta_excel: str) -> set:
    """Devuelve claves 'numero_documento|fecha' ya guardadas en el Excel.
    Funciona para cualquiera de las 3 plantillas, porque "tipo_linea",
    "numero_documento" y "fecha_documento" están siempre en las primeras
    columnas (la posición exacta se resuelve leyendo la fila 1 del Excel
    existente, no asumiendo una plantilla fija)."""
    ruta = Path(ruta_excel)
    if not ruta.exists() or ruta.suffix.lower() != ".xlsx":
        return set()
    try:
        wb = load_workbook(ruta, data_only=True)
    except Exception:
        return set()
    if NOMBRE_HOJA not in wb.sheetnames:
        return set()

    hoja = wb[NOMBRE_HOJA]

    # Detectar qué plantilla tiene el Excel existente comparando el
    # número de columnas con encabezado en la fila 1.
    ultima_col = hoja.max_column
    cols, claves, _ = _resolver_plantilla(None)
    for nombre_tipo, (plantilla, _dmap) in PLANTILLAS.items():
        if len(plantilla) == ultima_col:
            cols, claves = plantilla, [c[1] for c in plantilla]
            break

    idx_tipo  = claves.index("tipo_linea") + 1
    idx_num   = claves.index("numero_documento") + 1
    idx_fecha = claves.index("fecha_documento") + 1

    claves_existentes = set()
    for fila in hoja.iter_rows(min_row=3, values_only=True):
        if len(fila) < idx_tipo or fila[idx_tipo - 1] != "E":
            continue
        num   = fila[idx_num - 1]   if len(fila) >= idx_num   else None
        fecha = fila[idx_fecha - 1] if len(fila) >= idx_fecha else None
        if num is not None:
            fs = fecha.strftime("%Y-%m-%d") if isinstance(fecha, datetime) else str(fecha or "")
            claves_existentes.add(f"{num}|{fs}")
    return claves_existentes


# ── Construcción de filas ─────────────────────────────────────────────────────
def _construir_filas(documentos: list, claves: list, d_map: dict, tipo: str) -> list:
    """
    Construye las filas E y D para cada documento.

    d_map traduce claves internas (compartidas con la fila E) a claves
    semánticas de detalle (d_nombre_item, d_precio_unitario, etc.),
    que luego se resuelven a sus valores reales del PDF.

    Valores fijos por regla de negocio del portal EGEHID:
      - Fila E: ind_ingreso=1, tipo_pago=2, termino_pago=0
                cod_modificacion=3 SOLO en notas de crédito/débito
                fecha_expiracion_ncf = constante (Factura y Nota de Débito)
      - Fila D: d_ind_bien_servicio=2, d_ind_tasa_itbis=4
    Cualquier otro valor proviene exclusivamente del PDF.
    """
    # Invertir d_map para poder asignar: clave_interna ← valor_semántico
    inv_map = {v: k for k, v in d_map.items()}

    es_nota = tipo in ("nota_credito", "nota_debito")

    filas = []
    for doc in documentos:
        enc = doc["encabezado"]
        det = doc["detalle"]

        # ── Fila E ──────────────────────────────────────────────────────────
        e = {k: None for k in claves}
        e["tipo_linea"]           = "E"
        e["ncf"]                  = enc.get("ncf")
        e["fecha_documento"]      = enc.get("fecha_documento")
        e["numero_documento"]     = _a_entero(enc.get("numero_documento"))
        e["ind_ingreso"]          = 1      # Tipo de Ingresos — regla de negocio
        e["tipo_pago"]            = 2      # Tipo de Pago — regla de negocio
        e["fecha_vencimiento"]    = enc.get("fecha_vencimiento")
        e["moneda"]               = enc.get("moneda")
        e["tasa_cambio"]          = None
        e["rnc_cliente"]          = _a_entero(enc.get("rnc_cliente"))
        e["cliente"]              = enc.get("cliente")
        e["ncf_modificado"]       = enc.get("ncf_modificado")
        e["rnc_otro"]             = None
        e["fecha_ncf_modificado"] = enc.get("fecha_ncf_modificado")
        e["nota"]                 = enc.get("observaciones")
        # Fecha Expiración NCF: valor fijo, solo existe la columna en las
        # plantillas de Factura y Nota de Débito (no en Nota de Crédito).
        if "fecha_expiracion_ncf" in e:
            e["fecha_expiracion_ncf"] = FECHA_EXPIRACION_NCF
        # Código de Modificación: aplica únicamente a notas (crédito/débito),
        # que modifican otro comprobante. Las facturas lo dejan vacío.
        if "cod_modificacion" in e and es_nota:
            e["cod_modificacion"] = 3
        # Campos exclusivos de ND y Factura (None en NC por no existir la clave)
        if "termino_pago" in e:
            e["termino_pago"] = 0          # Término de Pago — regla de negocio
        if "razon_modificacion" in e:
            e["razon_modificacion"] = None
        e["_tipo"] = "E"
        filas.append(e)

        # ── Filas D ──────────────────────────────────────────────────────────
        for linea in det:
            d = {k: None for k in claves}
            d["tipo_linea"] = "D"
            d["ncf"]        = enc.get("ncf")

            # El PDF trae el monto TOTAL de la línea; el portal espera el
            # precio UNITARIO. Se deriva dividiendo entre la cantidad
            # (para cantidad = 1 ó ausente el valor no cambia).
            monto    = linea.get("monto")
            cantidad = linea.get("cantidad") or 1.0
            precio_unitario = (
                monto / cantidad if (monto is not None and cantidad) else monto
            )

            # Valores semánticos del detalle → extraídos del PDF o reglas fijas
            semanticos = {
                "d_nombre_item":       linea.get("nombre"),
                "d_desc_item":         linea.get("descripcion"),  # None si no hay
                "d_ind_bien_servicio": 2,                         # regla de negocio
                "d_ind_tasa_itbis":    4,                         # regla de negocio
                "d_cantidad":          linea.get("cantidad"),     # del PDF (o 1.0)
                "d_precio_unitario":   precio_unitario,           # monto / cantidad
                # Unidad de Medida se deja vacía a propósito: la plantilla
                # oficial no la trae y la DGII espera un código de catálogo,
                # no el texto libre del PDF ("kW", "kWh", …).
                "d_unidad_medida":     None,
            }

            # Asignar cada valor semántico a la clave interna correcta de esta plantilla
            for sem_key, valor in semanticos.items():
                clave_interna = inv_map.get(sem_key)
                if clave_interna and clave_interna in d:
                    d[clave_interna] = valor

            # termino_pago en fila E = 0 (regla de negocio).
            # En fila D, esa misma columna contiene "Precio Unitario Item"
            # (ya fue asignada arriba vía d_precio_unitario → inv_map),
            # así que NO se toca aquí para no sobreescribir el precio.

            d["_tipo"] = "D"
            filas.append(d)

    return filas


# ── Encabezados ───────────────────────────────────────────────────────────────
def _escribir_encabezados(hoja, cols):
    alin_c = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_num, clave, h1, h2, ancho in cols:
        c1 = hoja.cell(row=1, column=col_num, value=h1)
        c1.font      = FUENTE_HDR
        c1.fill      = GRIS_HDR
        c1.alignment = alin_c

        c2 = hoja.cell(row=2, column=col_num, value=h2)
        c2.font      = FUENTE_HDR
        c2.fill      = GRIS_HDR if h2 else GRIS_HDR_OSC
        c2.alignment = alin_c

        hoja.column_dimensions[get_column_letter(col_num)].width = ancho

    hoja.row_dimensions[1].height = 30
    hoja.row_dimensions[2].height = 30


# ── Escritura de datos ────────────────────────────────────────────────────────
def _ultima_fila_con_datos(hoja) -> int:
    """Devuelve el número de la última fila con al menos una celda con
    valor, ignorando filas completamente vacías que pudieran existir."""
    ultima = 2  # mínimo: después de los dos encabezados
    for fila in hoja.iter_rows():
        if any(c.value is not None for c in fila):
            ultima = fila[0].row
    return ultima


def _escribir_hoja(wb: Workbook, filas: list, cols: list, hoja_existente: bool = False):
    if NOMBRE_HOJA in wb.sheetnames:
        hoja = wb[NOMBRE_HOJA]
    else:
        hoja = wb.active
        hoja.title = NOMBRE_HOJA

    if not hoja_existente:
        _escribir_encabezados(hoja, cols)

    alin_izq  = Alignment(vertical="center",  wrap_text=False)
    alin_nota = Alignment(vertical="top",     wrap_text=False)

    # Calcular la fila de inicio UNA sola vez y llevar un contador; recalcular
    # la última fila en cada iteración era O(n²) sobre la hoja completa.
    num_fila = _ultima_fila_con_datos(hoja)

    for fila_dict in filas:
        num_fila += 1
        tipo     = fila_dict.get("_tipo", "E")

        for col_num, clave, h1, h2, *_ in cols:
            valor = fila_dict.get(clave)
            celda = hoja.cell(row=num_fila, column=col_num, value=valor)

            # Encabezado real de esta columna según la fila (E usa fila 1, D fila 2)
            encabezado_col = h1 if tipo == "E" else h2

            if isinstance(valor, datetime):
                celda.number_format = FORMATO_FECHA
            elif isinstance(valor, (int, float)) and _es_columna_monetaria(encabezado_col):
                # Importes (Precio Unitario, Monto…) como número con formato
                # de moneda, no como texto ni número plano.
                celda.number_format = FORMATO_NUM

            celda.alignment = alin_nota if clave == "nota" else alin_izq
            celda.font = FUENTE_E if tipo == "E" else FUENTE_D

            if tipo == "E":
                celda.fill = AZUL_E

    hoja.freeze_panes = "A3"


# ── Sobreescritura de duplicados ──────────────────────────────────────────────
def eliminar_filas_documento(hoja, claves: list, numero_documento: str, fecha_documento) -> None:
    """Elimina las filas E y D de un documento ya existente.
    Tolera filas en blanco dentro del bloque para no cortar la búsqueda."""
    idx_tipo  = claves.index("tipo_linea") + 1
    idx_num   = claves.index("numero_documento") + 1
    idx_fecha = claves.index("fecha_documento") + 1
    idx_ncf   = claves.index("ncf") + 1

    fecha_cmp = fecha_documento if isinstance(fecha_documento, datetime) else None

    filas_a_eliminar = []
    ncf_objetivo = None
    dentro_bloque = False

    for fila in hoja.iter_rows(min_row=3):
        tipo = fila[idx_tipo - 1].value
        num  = fila[idx_num - 1].value

        if tipo == "E":
            fecha = fila[idx_fecha - 1].value
            fecha_ok = (
                (fecha_cmp is not None and isinstance(fecha, datetime) and fecha.date() == fecha_cmp.date())
                or (fecha_cmp is None and fecha is None)
            )
            if str(num) == str(numero_documento) and fecha_ok:
                ncf_objetivo  = fila[idx_ncf - 1].value
                dentro_bloque = True
                filas_a_eliminar.append(fila[0].row)
            elif dentro_bloque:
                break

        elif tipo == "D" and dentro_bloque:
            ncf_fila = fila[idx_ncf - 1].value
            if ncf_fila == ncf_objetivo:
                filas_a_eliminar.append(fila[0].row)

        elif tipo is None and dentro_bloque:
            filas_a_eliminar.append(fila[0].row)

    for num_fila in sorted(filas_a_eliminar, reverse=True):
        hoja.delete_rows(num_fila)


# ── Punto de entrada público ──────────────────────────────────────────────────
def generar_excel(documentos: list, ruta_salida: str, tipo_documento: str | None = None):
    """
    Genera (o actualiza) el Excel de salida usando la plantilla correcta
    según `tipo_documento`: "factura" | "nota_credito" | "nota_debito".

    Si `tipo_documento` es None, se intenta inferir del primer documento
    de la lista (campo encabezado["tipo_documento"]); si tampoco está
    disponible, se usa la plantilla de Nota de Crédito por compatibilidad
    con versiones anteriores del proyecto.
    """
    if tipo_documento is None and documentos:
        tipo_documento = documentos[0]["encabezado"].get("tipo_documento")

    cols, claves, d_map = _resolver_plantilla(tipo_documento)
    tipo_efectivo = tipo_documento if tipo_documento in PLANTILLAS else TIPO_POR_DEFECTO

    ruta = Path(ruta_salida)
    hoja_existente = False

    if ruta.exists() and ruta.suffix.lower() == ".xlsx":
        try:
            wb = load_workbook(ruta)
            hoja_existente = True
        except Exception:
            wb = Workbook()
    else:
        wb = Workbook()

    if hoja_existente and NOMBRE_HOJA in wb.sheetnames:
        hoja = wb[NOMBRE_HOJA]
        for doc in documentos:
            if doc.get("_sobreescribir"):
                enc = doc["encabezado"]
                eliminar_filas_documento(
                    hoja, claves,
                    numero_documento=enc.get("numero_documento"),
                    fecha_documento=enc.get("fecha_documento"),
                )

    filas = _construir_filas(documentos, claves, d_map, tipo_efectivo)
    _escribir_hoja(wb, filas, cols, hoja_existente=hoja_existente)
    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(ruta))


# ── Test rápido ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import pdfplumber, sys
    sys.path.insert(0, str(Path(__file__).parent))
    from extractor_encabezado import extraer_encabezado
    from extractor_detalle import extraer_detalle

    ruta_pdf = Path("../muestras/800001168.pdf")
    textos, palabras_todas, offset = [], [], 0.0
    with pdfplumber.open(ruta_pdf) as pdf:
        for p in pdf.pages:
            textos.append(p.extract_text() or "")
            for w in p.extract_words():
                c = dict(w); c["top"] += offset; palabras_todas.append(c)
            ws = p.extract_words()
            if ws: offset += max(w["bottom"] for w in ws) + 10

    enc = extraer_encabezado("\n".join(textos))
    det = extraer_detalle(palabras_todas, numero_documento=enc.get("numero_documento"))
    generar_excel([{"encabezado": enc, "detalle": det}], "../salida/resultado.xlsx")
    print("Excel generado. Tipo detectado:", enc.get("tipo_documento"))